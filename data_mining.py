import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from tqdm import tqdm
from sklearn.metrics import (
    r2_score,
    mean_absolute_percentage_error,
    mean_squared_error,
    mean_absolute_error,
)

# ===== Excel writer (đúng format như hình) =====
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def export_metrics_excel_like_template(
    combined_df,                 # index = lead_time, columns = "<model>|<metric>"
    out_xlsx_path,
    models,                      # ví dụ: ["Our model", "S2S"]
    metrics=("MAE", "R2", "RMSE", "Corr"),
    round_digits=4,
):
    """
    Xuất Excel dạng:
    Leadtime | MAE(Our model,S2S) | R2(...) | RMSE(...) | Corr(...)
    Header 2 tầng và merge cell giống template bạn gửi.
    """
    os.makedirs(os.path.dirname(out_xlsx_path) or ".", exist_ok=True)

    # Làm tròn dữ liệu
    df = combined_df.copy()
    df = df.round(round_digits)

    wb = Workbook()
    ws = wb.active
    ws.title = "metrics"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    # Header A1:A2
    ws["A1"] = "Leadtime"
    ws["A1"].alignment = center
    ws["A1"].font = bold
    ws.merge_cells("A1:A2")

    # Header nhóm metrics
    start_col = 2  # cột B
    n_models = len(models)

    for mi, metric in enumerate(metrics):
        c0 = start_col + mi * n_models
        c1 = c0 + n_models - 1

        ws.cell(row=1, column=c0, value=metric).alignment = center
        ws.cell(row=1, column=c0).font = bold
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)

        for k, model in enumerate(models):
            ws.cell(row=2, column=c0 + k, value=model).alignment = center
            ws.cell(row=2, column=c0 + k).font = bold

    # Data rows (lead_time)
    lead_times = list(df.index)
    for r, lt in enumerate(lead_times, start=3):
        ws.cell(row=r, column=1, value=int(lt)).alignment = center

        for mi, metric in enumerate(metrics):
            c0 = start_col + mi * n_models
            # map Corr -> COR (đúng key trong combined_df)
            key_metric = "COR" if metric.lower() == "corr" else metric

            for k, model in enumerate(models):
                key = f"{model}|{key_metric}"
                val = df.loc[lt, key] if key in df.columns else np.nan

                cell = ws.cell(
                    row=r,
                    column=c0 + k,
                    value=None if (val is None or (isinstance(val, float) and np.isnan(val))) else float(val),
                )
                cell.alignment = center
                cell.number_format = "0." + ("0" * round_digits)

    # chỉnh độ rộng cột
    total_cols = 1 + len(metrics) * n_models
    for col in range(1, total_cols + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 12 if col == 1 else 16

    wb.save(out_xlsx_path)
    print(f"[DONE] Saved Excel: {out_xlsx_path}")


# ===== Convert W&B CSV (giữ nguyên logic của bạn) =====
def convert_wandb_csv_subseasonal(input_csv, output_csv, lead_begin, lead_end):
    station_csv = "/mnt/disk3/longnd/env_data/Gauge_thay_Tan/Final_Data_Region_1.csv"
    time_csv = "/mnt/disk3/tunm/Subseasonal_Forecasting/data3/data6789_reg_1_seed52_new_all/test.csv"

    df_pred = pd.read_csv(input_csv)
    df_station = pd.read_csv(station_csv)
    df_time = pd.read_csv(time_csv)

    stations = df_station["Station"].unique()
    num_stations = len(stations)
    num_lead_time = lead_end - lead_begin + 1

    num_days_possible = 111361 // (num_stations * num_lead_time)
    num_days = int(num_days_possible)

    total_values_used = num_days * num_lead_time * num_stations
    print(f"Số ngày: {num_days}, Tổng giá trị sử dụng: {total_values_used}")
    assert total_values_used <= 200000, "Số giá trị vượt quá 200,000"

    data = []
    for day_idx in tqdm(range(num_days * num_lead_time), desc="Processing"):
        lead_time = df_time.iloc[day_idx]["leadTime"]
        year = df_time.iloc[day_idx]["year"]
        month = df_time.iloc[day_idx]["month"]
        day = df_time.iloc[day_idx]["day"]

        start_idx = day_idx * num_stations
        end_idx = start_idx + num_stations
        group_data = df_pred.iloc[start_idx:end_idx]

        for i, station in enumerate(stations):
            if i >= len(group_data):
                print(f"[Warning] day_idx {day_idx}, station_idx {i} out of bounds with group_data len = {len(group_data)}")
                continue

            row = {
                "Prediction": group_data.iloc[i]["Prediction"],
                "Groundtruth": group_data.iloc[i]["Groundtruth"],
                "station": station,
                "lead_time": lead_time,
                "year": year,
                "month": month,
                "day": day,
            }
            data.append(row)

    df_output = pd.DataFrame(data)
    os.makedirs(os.path.dirname(output_csv) or ".", exist_ok=True)
    df_output.to_csv(output_csv, index=False)
    print(f"[DONE] Created {output_csv}")


# ===== Metrics =====
def cal_acc(y_prd, y_grt):
    mae = mean_absolute_error(y_grt, y_prd)
    mse = mean_squared_error(y_grt, y_prd)
    mape = mean_absolute_percentage_error(y_grt, y_prd)
    rmse = np.sqrt(mse)
    corr = np.corrcoef(np.reshape(y_grt, (-1)), np.reshape(y_prd, (-1)))[0][1]
    r2 = r2_score(y_grt, y_prd)
    return mae, mse, mape, rmse, r2, corr


def calculate_metrics_per_station(csv_file, start_lead_time=1, end_lead_time=46):
    print(f"Đang tải dữ liệu từ: {csv_file}")
    df = pd.read_csv(csv_file)

    df_filtered = df[(df["lead_time"] >= start_lead_time) & (df["lead_time"] <= end_lead_time)]
    if df_filtered.empty:
        print("Không tìm thấy dữ liệu trong khoảng lead time đã chỉ định.")
        return None

    mae, mse, mape, rmse, r2, corr = cal_acc(df["Prediction"], df["Groundtruth"])
    print(f"Global: MAE: {mae}, R2: {r2}, corr: {corr}")

    def calculate_group_metrics(group):
        predictions = group["Prediction"]
        groundtruths = group["Groundtruth"]
        if len(predictions) < 2:
            return pd.Series({"MAE": np.nan, "R2": np.nan, "Data Points": len(predictions)})
        mae_ = mean_absolute_error(groundtruths, predictions)
        r2_ = r2_score(groundtruths, predictions)
        return pd.Series({"MAE": mae_, "R2": r2_, "Data Points": len(predictions)})

    station_metrics = df_filtered.groupby("station").apply(calculate_group_metrics)
    print(station_metrics.to_string())

    print("\nTổng kết chung:")
    print(f"  - MAE trung bình: {station_metrics['MAE'].mean():.4f}")
    print(f"  - R² trung bình:  {station_metrics['R2'].mean():.4f}")

    return station_metrics


# ===== Plot comparison + export CSV + export XLSX (ALL METRICS) =====
def calculate_and_plot_comparison(
    start_lead_time=1,
    end_lead_time=46,
    metric_to_plot="R2",  # 'R2' | 'MAE' | 'COR' | 'RMSE'
    save_plot_path="metric_comparison.png",
    save_csv_path="metric_comparison.csv",
    save_all_metrics_csv_path="metric_comparison_all_metrics.csv",
    save_all_metrics_xlsx_path="metric_comparison_all_metrics.xlsx",
    round_digits=4,
):
    all_results = {}

    csv_configs = {
        "S2S": {
            "file": "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/ecmwf-final.csv",
            "marker": "o",
            "linestyle": "-",
            "color": "#1f77b4",   # xanh dương
        },
        "Our model": {
            "file": "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/vit-final.csv",
            "marker": "s",
            "linestyle": "--",
            "color": "#d62728",   # đỏ
        },
        "CNN-LSTM": {
            "file": "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/cnn-lstm-final.csv",
            "marker": "^",
            "linestyle": "-.",
            "color": "#2ca02c",   # xanh lá
        },
        # "Without GsMAP": {
        #     "file": "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/vit_gsmap-final.csv",
        #     "marker": "D",
        #     "linestyle": ":",
        #     "color": "#ff7f0e",   # cam
        # },
        # "Without Addition LT": {
        #     "file": "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/vit_lt-final.csv",
        #     "marker": "P",
        #     "linestyle": (0, (3, 1, 1, 1)),
        #     "color": "#9467bd",   # tím
        # },
    }

    # --- TÍNH METRICS ---
    for model_name, config in csv_configs.items():
        csv_file = config["file"]
        print("-" * 60)
        print(f">>> Xử lý mô hình: {model_name}")
        print(f"    File: {csv_file}")

        try:
            df = pd.read_csv(csv_file)
        except FileNotFoundError:
            print(f"    Lỗi: Không tìm thấy file '{csv_file}'. Bỏ qua.")
            all_results[model_name] = pd.DataFrame()
            continue

        model_run_results = []
        for lead_time in range(start_lead_time, end_lead_time + 1):
            df_lt = df[df["lead_time"] == lead_time]
            if df_lt.empty:
                continue

            # ===== GLOBAL cho RMSE & COR: gộp tất cả station =====
            df_global = df_lt.dropna(subset=["Prediction", "Groundtruth"])
            pred_g = df_global["Prediction"].to_numpy()
            gt_g = df_global["Groundtruth"].to_numpy()

            if len(pred_g) > 1:
                global_rmse = float(np.sqrt(mean_squared_error(gt_g, pred_g)))
                global_cor = float(np.corrcoef(gt_g, pred_g)[0, 1])
            else:
                global_rmse = np.nan
                global_cor = np.nan

            # ===== MAE & R2: theo station rồi trung bình (GIỮ NGUYÊN) =====
            stations = df_lt["station"].unique()
            station_maes, station_r2s = [], []

            for station in stations:
                df_station = df_lt[df_lt["station"] == station].dropna(
                    subset=["Prediction", "Groundtruth"]
                )
                predictions = df_station["Prediction"]
                groundtruths = df_station["Groundtruth"]

                if len(predictions) > 1:
                    station_maes.append(mean_absolute_error(groundtruths, predictions))
                    station_r2s.append(r2_score(groundtruths, predictions))

            if len(station_maes) > 0:
                model_run_results.append(
                    {
                        "lead_time": lead_time,
                        "Average MAE": float(np.mean(station_maes)),
                        "Average R2": float(np.mean(station_r2s)),
                        "Average COR": global_cor,     # GLOBAL
                        "Average RMSE": global_rmse,   # GLOBAL
                    }
                )

        all_results[model_name] = pd.DataFrame(model_run_results)
        print(f">>> Hoàn thành: {model_name}")

    # --- (1) CSV 1 metric như cũ ---
    column_to_plot = f"Average {metric_to_plot}"
    lead_times = list(range(start_lead_time, end_lead_time + 1))
    comparison_df = pd.DataFrame({"lead_time": lead_times}).set_index("lead_time")

    for model_name, df_metrics in all_results.items():
        if df_metrics is None or df_metrics.empty or column_to_plot not in df_metrics.columns:
            print(f"    Cảnh báo: Thiếu dữ liệu '{column_to_plot}' cho {model_name} -> để NaN")
            comparison_df[model_name] = np.nan
            continue

        s = df_metrics.set_index("lead_time")[column_to_plot]
        comparison_df[model_name] = s.reindex(lead_times)

    comparison_df = comparison_df.round(round_digits)

    os.makedirs(os.path.dirname(save_csv_path) or ".", exist_ok=True)
    comparison_df.to_csv(save_csv_path, index=True)
    print(f"[DONE] Saved CSV (1 metric): {save_csv_path}")

    # --- (2) CSV + XLSX TỔNG 4 METRICS ---
    metrics_list = ["MAE", "R2", "RMSE", "COR"]  # internal keys
    combined_df = pd.DataFrame(index=lead_times)

    for model_name, df_metrics in all_results.items():
        if df_metrics is None or df_metrics.empty:
            for m in metrics_list:
                combined_df[f"{model_name}|{m}"] = np.nan
            continue

        df_m = df_metrics.set_index("lead_time")
        for m in metrics_list:
            col = f"Average {m}"
            combined_df[f"{model_name}|{m}"] = (
                df_m[col].reindex(lead_times) if col in df_m.columns else np.nan
            )

    combined_df.index.name = "lead_time"
    combined_df = combined_df.round(round_digits)

    os.makedirs(os.path.dirname(save_all_metrics_csv_path) or ".", exist_ok=True)
    combined_df.to_csv(save_all_metrics_csv_path, index=True)
    print(f"[DONE] Saved CSV (all metrics): {save_all_metrics_csv_path}")

    # Export XLSX (format như hình)
    models_for_excel = ["Our model", "S2S", "CNN-LSTM"]
    export_metrics_excel_like_template(
        combined_df,
        out_xlsx_path=save_all_metrics_xlsx_path,
        models=models_for_excel,
        metrics=("MAE", "R2", "RMSE", "Corr"),
        round_digits=round_digits,
    )

    # --- VẼ BIỂU ĐỒ ---
    print(f"Vẽ biểu đồ so sánh: {column_to_plot}")
    plt.figure(figsize=(14, 8))

    any_plotted = False
    for model_name, config in csv_configs.items():
        if model_name not in comparison_df.columns:
            continue
        y = comparison_df[model_name].values
        if np.all(np.isnan(y)):
            continue

        plt.plot(
            comparison_df.index.values,
            y,
            label=model_name,
            marker=config["marker"],
            linestyle=config["linestyle"],
            color=config["color"],
        )
        any_plotted = True

    if not any_plotted:
        print("Không có dữ liệu để vẽ.")
        plt.close()
        return comparison_df, combined_df

    plt.title(f"Comparison of Average {metric_to_plot} Score by Lead Time", fontsize=16)
    plt.xlabel("Lead Time (days)", fontsize=12)
    plt.ylabel(f"Average {metric_to_plot} Score", fontsize=12)
    plt.legend(fontsize=12, loc="best")
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)

    if metric_to_plot in ["R2", "COR"]:
        plt.axhline(0, color="red", linestyle=":", linewidth=1.2)

    plt.xticks(np.arange(start_lead_time, end_lead_time + 1, 2))
    plt.xlim(start_lead_time - 0.5, end_lead_time + 0.5)

    os.makedirs(os.path.dirname(save_plot_path) or ".", exist_ok=True)
    plt.tight_layout()
    plt.savefig(save_plot_path)
    print(f"[DONE] Saved plot: {save_plot_path}")
    plt.show()
    plt.close()

    return comparison_df, combined_df


# ===== (giữ nguyên) plot_test của bạn – không đụng gì =====
def plot_test(csv_file, csv_file_v2, lead_time_value=7, model1="VIT", model2="S2S"):
    saved_dir = f"/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/plot/leadtime{lead_time_value}"
    os.makedirs(saved_dir, exist_ok=True)

    df = pd.read_csv(csv_file)
    df_ver2 = pd.read_csv(csv_file_v2)

    filtered_df = df[(df["lead_time"] == lead_time_value) & (1)]
    filtered_df_ver2 = df_ver2[(df_ver2["lead_time"] == lead_time_value) & (1)]

    stations = filtered_df["station"].unique()

    for station_value in stations:
        station_df = filtered_df[filtered_df["station"] == station_value].copy()
        station_df_ver2 = filtered_df_ver2[filtered_df_ver2["station"] == station_value].copy()

        def make_date_or_feb28(df_, y="year", m="month", d="day", out="date"):
            dt = pd.to_datetime(df_[[y, m, d]], errors="coerce")
            feb28 = pd.to_datetime({"year": df_[y], "month": 2, "day": 28}, errors="coerce")
            df_[out] = dt.fillna(feb28)
            return df_

        station_df = make_date_or_feb28(station_df, "year", "month", "day", "date")
        station_df_ver2 = make_date_or_feb28(station_df_ver2, "year", "month", "day", "date")

        station_df = station_df.sort_values("date")
        station_df_ver2 = station_df_ver2.sort_values("date")

        indices = range(len(station_df))

        predictions = station_df["Prediction"]
        groundtruths = station_df["Groundtruth"]
        predictions_ver2 = station_df_ver2["Prediction"]

        stran_mae, _, _, stran_rmse, stran_r2, stran_corr = cal_acc(predictions, groundtruths)
        ecmwf_mae, _, _, ecmwf_rmse, ecmwf_r2, ecmwf_corr = cal_acc(predictions_ver2, groundtruths)

        text_str = (
            f"{model1}:\n"
            f"  MAE: {stran_mae:.4f}\n"
            f"  RMSE: {stran_rmse:.4f}\n"
            f"  R²: {stran_r2:.4f}\n"
            f"  Corr: {stran_corr:.4f}\n\n"
            f"{model2}:\n"
            f"  MAE: {ecmwf_mae:.4f}\n"
            f"  RMSE: {ecmwf_rmse:.4f}\n"
            f"  R²: {ecmwf_r2:.4f}\n"
            f"  Corr: {ecmwf_corr:.4f}\n\n"
        )

        plt.figure(figsize=(12, 6))
        plt.plot(indices, groundtruths, color="red", markersize=1, marker="d", linestyle="-", alpha=0.75,
                 label="Groundtruth (Gauge)", linewidth=0.7)
        plt.plot(indices, predictions, color="blue", markersize=1, marker="o", linestyle="--", alpha=0.9,
                 label=model1, linewidth=0.6)
        # plt.plot(indices, predictions_ver2, color="green", markersize=1, marker="s", linestyle=":", alpha=0.9,
        #          label=model2, linewidth=0.8)

        props = dict(boxstyle="round", facecolor="white", edgecolor="black", lw=1)
        ax = plt.gca()
        # ax.text(0.95, 0.95, text_str, transform=ax.transAxes, fontsize=10, verticalalignment="top", bbox=props)

        plt.title(f"{model1} and {model2} Comparison for Station {station_value} - Lead Time = {lead_time_value}", fontsize=14)
        plt.xlabel("Index (Days)", fontsize=12)
        plt.ylabel("Rainfall (mm)", fontsize=12)
        plt.legend(loc="upper left")
        plt.grid(False)
        plt.tight_layout()

        plt.savefig(f"{saved_dir}/prediction_comparison_station_{station_value}_lead_time_{lead_time_value}.png")
        plt.close()
        if station_value == 13 and lead_time_value == 7:
            station_df["Groundtruth"].to_excel("temp.xlsx", index=False)


if __name__ == "__main__":
    vit_file = "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/vit.csv"
    ecmwf_file = "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/ecmwf.csv"
    cnn_lstm_file = "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/cnn-lstm.csv"
    vit_lt_file = "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/vit_lt.csv"
    vit_gsmap_file = "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/vit_gsmap.csv"

    vit_file_final = "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/vit-final.csv"
    ecmwf_file_final = "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/ecmwf-final.csv"
    cnn_lstm_file_final = "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/cnn-lstm-final.csv"
    vit_lt_file_final = "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/vit_lt-final.csv"
    vit_gsmap_file_final = "/mnt/disk3/tunm/Subseasonal_Forecasting/results/Strans/0412/vit_gsmap-final.csv"

    begin_lead = 7
    end_lead = 46

    # convert sang format final
    # convert_wandb_csv_subseasonal(vit_file, vit_file_final, begin_lead, end_lead)
    # convert_wandb_csv_subseasonal(ecmwf_file, ecmwf_file_final, begin_lead, end_lead)
    # convert_wandb_csv_subseasonal(cnn_lstm_file, cnn_lstm_file_final, begin_lead, end_lead)
    # convert_wandb_csv_subseasonal(vit_lt_file, vit_lt_file_final, begin_lead, end_lead)
    # convert_wandb_csv_subseasonal(vit_gsmap_file, vit_gsmap_file_final, begin_lead, end_lead)
    model1 = "CNN_LSTM"
    model2 = "S2S"

    for lead in range(begin_lead, end_lead+1):
        plot_test(ecmwf_file_final, ecmwf_file_final, lead_time_value=lead, model1=model1, model2=model2)

    # _ = calculate_metrics_per_station(vit_file_final, start_lead_time=begin_lead, end_lead_time=end_lead)

    # 4 plots như bạn đang làm — mỗi lần đều xuất (ghi đè) all_metrics.csv và all_metrics.xlsx
    # for metric in ["R2", "RMSE", "COR", "MAE"]:
    #     calculate_and_plot_comparison(
    #         start_lead_time=begin_lead,
    #         end_lead_time=end_lead,
    #         metric_to_plot=metric,
    #         save_plot_path=f"fig/{metric.lower()}_strans_vs_ecmwf_comparison.png",
    #         save_csv_path=f"fig/{metric}_comparison.csv",
    #         save_all_metrics_csv_path="fig/all_metrics_comparison.csv",
    #         save_all_metrics_xlsx_path="fig/all_metrics_comparison.xlsx",
    #         round_digits=2,
    #     )
